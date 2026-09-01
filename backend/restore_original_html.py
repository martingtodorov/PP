"""Restore the original Shopify body HTML (products, collections, pages, articles).

`clean_imported_html.py` deleted the leading <h1> ("Какво представлява X?") instead of demoting it.
This re-reads the Matrixify export and keeps the heading, only downgrading <h1> to <h2> so the page
still has a single H1 (the product title).
"""
import logging
import os
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))

XLSX = ROOT / "data" / "matrixify-export.xlsx"
log = logging.getLogger("restore_html")


def demote_h1(html: str) -> str:
    html = re.sub(r"<h1(\s[^>]*)?>", lambda m: f"<h2{m.group(1) or ''}>", html, flags=re.I)
    return re.sub(r"</h1>", "</h2>", html, flags=re.I)


def sheet_rows(wb, name: str):
    if name not in wb.sheetnames:
        return
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    for row in rows:
        yield dict(zip(header, row))


def run() -> dict:
    db = MongoClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    out = {}

    # Products
    bodies = {}
    for r in sheet_rows(wb, "Products"):
        handle, body = r.get("Handle"), r.get("Body HTML")
        if handle and body and handle not in bodies:
            bodies[handle] = body
    n = 0
    for handle, body in bodies.items():
        res = db.products.update_one({"handle": handle}, {"$set": {"description": demote_h1(str(body).strip())}})
        n += res.modified_count
    out["products"] = {"restored": n, "in_export": len(bodies)}

    # Collections
    n = 0
    for sheet in ("Smart Collections", "Custom Collections"):
        for r in sheet_rows(wb, sheet):
            handle, body = r.get("Handle"), r.get("Body HTML")
            if handle and body:
                if handle == "all-peptides":
                    handle = "2all-the-peptides-1"
                n += db.collections_cat.update_one(
                    {"handle": handle}, {"$set": {"description": demote_h1(str(body).strip())}}
                ).modified_count
    out["collections"] = n

    # Pages (Bulgarian source only)
    slug_map = {"contacts": "contact-1", "partners": "become-a-distributor", "about": "about-1",
                "terms-of-service": "terms-conditions", "shipping-policy": "delivery-and-payment",
                "what-are-peptides": "какво-са-пептиди"}
    n = 0
    for r in sheet_rows(wb, "Pages"):
        handle, body = r.get("Handle"), r.get("Body HTML")
        if not handle or not body:
            continue
        slug = slug_map.get(handle, handle)
        n += db.pages.update_one({"slug": slug, "locale": "bg"},
                                 {"$set": {"html": demote_h1(str(body).strip())}}).modified_count
    out["pages"] = n

    # Articles
    n = 0
    for r in sheet_rows(wb, "Blog Posts"):
        handle, body = r.get("Handle"), r.get("Body HTML")
        if handle and body:
            n += db.articles.update_one({"handle": handle},
                                        {"$set": {"body": demote_h1(str(body).strip())}}).modified_count
    out["articles"] = n
    return out


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    import json
    print(json.dumps(run(), ensure_ascii=False, indent=1))
