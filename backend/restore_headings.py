"""Put the Shopify body heading back at the top of product and collection descriptions.

Earlier imports dropped the body's own <h1> because the page already had a title. The live
purepeptide.bg shows that heading as the H1 ("Какво е Ретатрутид?", "Пептиди, изследвани за
отслабване и метаболизъм") and the owner wants the same, so the heading is restored from the
Matrixify export without touching anything else (uploaded pictures, prices, translations).
Idempotent — runs at every startup.
"""
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, Optional

from starlette.concurrency import run_in_threadpool

log = logging.getLogger("purepeptide.headings")
BUNDLED_XLSX = Path(__file__).parent / "data" / "matrixify-export.xlsx"
_H1 = re.compile(r"^\s*<h1[^>]*>(.*?)</h1>", re.I | re.S)
SHEETS = {"products": ("Products",), "collections": ("Smart Collections", "Custom Collections")}


def _norm(text: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", re.sub(r"<[^>]+>", "", text or "").lower())


def body_headings(source, sheets=SHEETS["products"]) -> Dict[str, str]:
    """Handle -> raw leading <h1> text of the Shopify body, for the given sheets."""
    import openpyxl

    wb = openpyxl.load_workbook(source, read_only=True)
    out: Dict[str, str] = {}
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        rows = wb[name].iter_rows(values_only=True)
        header = list(next(rows))
        ih, ib = header.index("Handle"), header.index("Body HTML")
        for row in rows:
            handle, body = row[ih], row[ib]
            if not handle or not body or str(handle) in out:
                continue
            m = _H1.match(str(body))
            if m:
                heading = re.sub(r"<[^>]+>", "", m.group(1)).strip()
                if heading:
                    out[str(handle).strip()] = heading
    return out


async def _latest_export(db, storage) -> Optional[bytes]:
    job = await db.import_jobs.find_one({"type": "matrixify", "status": "completed"},
                                        {"_id": 0, "storage_path": 1}, sort=[("at", -1)])
    if not job or not job.get("storage_path"):
        return None
    try:
        blob, _ = await run_in_threadpool(storage.get_object, job["storage_path"])
        return blob
    except Exception:
        return None


async def _restore(coll, headings: Dict[str, str]) -> int:
    fixed = 0
    for handle, heading in headings.items():
        doc = await coll.find_one({"handle": handle}, {"_id": 1, "description": 1})
        if not doc:
            continue
        desc = doc.get("description") or ""
        if _H1.match(desc) or _norm(heading) in _norm(desc[:400]):
            continue
        await coll.update_one({"_id": doc["_id"]}, {"$set": {"description": f"<h1>{heading}</h1>\n{desc}"}})
        fixed += 1
    return fixed


async def restore_headings(db, storage) -> Dict[str, int]:
    blob = await _latest_export(db, storage)
    if not blob and not BUNDLED_XLSX.exists():
        return {}
    source = BytesIO(blob) if blob else BUNDLED_XLSX
    result = {}
    for kind, coll in (("products", db.products), ("collections", db.collections_cat)):
        if isinstance(source, BytesIO):
            source.seek(0)
        headings = await run_in_threadpool(body_headings, source, SHEETS[kind])
        result[kind] = await _restore(coll, headings)
    if any(result.values()):
        log.info("Restored the body heading: %s", result)
    return result
