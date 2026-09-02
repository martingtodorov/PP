"""One-off importer for the real purepeptide.bg Matrixify export.

Reads /app/backend/data/matrixify-export.xlsx and replaces the demo catalog with the
real products, collections, pages, blog posts, redirects, discounts, customers and
orders. Shopify CDN images are downloaded into our own object storage.

Usage: python matrixify_import.py [--skip-images] [--only products,collections,...]
"""

import hashlib
import html as html_lib
import logging
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import requests
from dotenv import load_dotenv
from pymongo import MongoClient

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))

import storage  # noqa: E402
from currency import rate_for  # noqa: E402
from links_map import link_key_for  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("import")

XLSX = ROOT / "data" / "matrixify-export.xlsx"
db = MongoClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]

ALL_HANDLE = "2all-the-peptides-1"   # must match server.ALL_COLLECTION, otherwise the catch-all 404s
SHOPIFY_ALL = "2all-the-peptides-1"

# Shopify page handle -> our static page slug
PAGE_MAP = {
    "какво-са-пептиди": "what-are-peptides",
    "chemical-analysis": "chemical-analysis",
    "faq": "faq",
    "contact-1": "contacts",
    "become-a-distributor": "partners",
    "about-1": "about",
    "terms-conditions": "terms-of-service",
    "delivery-and-payment": "shipping-policy",
    "cookies": "cookies",
    "scientific-literature": "scientific-literature",
    "data-sharing-opt-out": "privacy-policy",
}

# Preferred storefront order / menu labels for the real collections
COLLECTION_META = {
    ALL_HANDLE: ("Всички пептиди", 0),
    "metabolic-studies": ("Пептиди за Отслабване", 1),
    "studies-on-healing": ("Пептиди за Възстановяване", 2),
    "secretagogues": ("Пептиди за Мускули", 3),
    "longevity-and-more": ("Пептиди за Кожа", 4),
    "melanin-i-libido": ("Пептиди за Либидо и Меланин", 5),
    "immunology": ("Пептиди за Имунитет", 6),
}

MIME = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp", "gif": "image/gif"}


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_dt(value: Any) -> str:
    if not value:
        return now_utc()
    if isinstance(value, datetime):
        return value.replace(tzinfo=value.tzinfo or timezone.utc).isoformat()
    try:
        return datetime.strptime(str(value)[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc).isoformat()
    except ValueError:
        return now_utc()


def num(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return default


def sheet(name: str) -> List[Dict[str, Any]]:
    ws = WB[name]
    rows = ws.iter_rows(values_only=True)
    hdr = [h for h in next(rows)]
    out = []
    for r in rows:
        out.append({h: r[i] for i, h in enumerate(hdr) if h})
    return out


def group_by(rows: List[Dict[str, Any]], key: str) -> Dict[str, List[Dict[str, Any]]]:
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        k = r.get(key)
        if k is None:
            continue
        groups.setdefault(str(k), []).append(r)
    return groups


# ---------- images ----------
SKIP_IMAGES = "--skip-images" in sys.argv


def _stored_ok(path: Optional[str]) -> bool:
    """A cached image_map entry is only usable while the file is really still in storage.

    A deploy to a fresh server (or a wiped media directory) leaves the mapping behind, and trusting
    it silently published /api/files URLs that 404 for every product.
    """
    if not path:
        return False
    try:
        storage.get_object(path)
        return True
    except Exception:
        return False


def store_image(url: Optional[str]) -> Optional[str]:
    """Download a remote image once and return our own /api/files/... URL."""
    if not url or not str(url).startswith("http"):
        return url
    url = str(url)
    if SKIP_IMAGES:
        return url
    # Shopify appends a ?v=<version> query that changes on every edit; keying the stored object on
    # the bare URL keeps re-imports pointing at the same file instead of minting orphan paths.
    norm = url.split("?")[0]
    cached = db.image_map.find_one({"$or": [{"key": norm}, {"src": url}]})
    if cached and _stored_ok(cached.get("path")):
        return cached["url"]
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
        data = resp.content
    except Exception as ex:
        log.warning("image download failed %s (%s)", url[:90], ex)
        return url
    base = url.split("?")[0].rsplit("/", 1)[-1] or "image"
    base = re.sub(r"[^A-Za-z0-9._-]", "-", base)
    ext = base.rsplit(".", 1)[-1].lower() if "." in base else "png"
    content_type = MIME.get(ext, resp.headers.get("Content-Type", "image/png").split(";")[0])
    path = f"import/{hashlib.sha1(norm.encode()).hexdigest()[:12]}-{base}"
    try:
        result = storage.put_object(path, data, content_type)
    except Exception as ex:
        log.warning("image upload failed %s (%s)", base, ex)
        return url
    stored = result.get("path", path)
    db.files.update_one(
        {"storage_path": stored},
        {"$set": {"storage_path": stored, "original_filename": base, "content_type": content_type,
                  "size": len(data), "is_deleted": False, "uploaded_by": "matrixify-import"},
         "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now_utc()}},
        upsert=True,
    )
    our_url = f"/api/files/{stored}"
    db.image_map.update_one(
        {"key": norm},
        {"$set": {"key": norm, "src": url, "path": stored, "url": our_url},
         "$setOnInsert": {"created_at": now_utc()}},
        upsert=True,
    )
    log.info("stored image %s (%d kB)", base, len(data) // 1024)
    return our_url


IMG_SRC_RE = re.compile(r'src="(https://cdn\.shopify\.com/[^"]+)"')


def rewrite_body_images(body: Optional[str]) -> str:
    if not body:
        return ""
    return IMG_SRC_RE.sub(lambda m: f'src="{store_image(m.group(1))}"', body)


SPEC_PATTERNS = {
    "cas": re.compile(r"CAS\s*#?\s*:?\s*([0-9]{2,7}-[0-9]{2}-[0-9])"),
    "formula": re.compile(r"(?:Формула|Formula)\s*:?\s*([A-Za-z0-9]+)"),
    "mw": re.compile(r"(?:М\.?\s?Т|M\.?W|Молекулна маса|Molecular weight)[^0-9]{0,12}([0-9][0-9.,]*\s*(?:g/mol|г/мол)?)"),
    "purity": re.compile(r"(>\s?9[0-9](?:[.,][0-9])?\s?%)"),
}


def extract_specs(body: str) -> Dict[str, str]:
    text = html_lib.unescape(re.sub(r"<[^>]+>", " ", body or ""))
    specs = {}
    for key, pattern in SPEC_PATTERNS.items():
        m = pattern.search(text)
        if m:
            specs[key] = m.group(1).strip()
    return specs


def strip_html(value: Optional[str], limit: int = 200) -> str:
    text = html_lib.unescape(re.sub(r"<[^>]+>", " ", value or ""))
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def norm_text(value: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", (value or "").lower())


def meta(row: Dict[str, Any], key: str) -> str:
    """A Shopify metafield export can carry either column type — read both."""
    for col in (f"Metafield: {key} [string]", f"Metafield: {key} [single_line_text_field]"):
        value = str(row.get(col) or "").strip()
        if value:
            return value
    return ""


def seo_pair(row: Dict[str, Any], title: str, body: str, limit: int = 160) -> Dict[str, str]:
    """Every record must ship a meta title and description — fall back to the title / body text."""
    return {
        "seo_title": meta(row, "title_tag") or (title or "").strip(),
        "seo_description": (meta(row, "description_tag") or strip_html(body, limit)
                            or (title or "").strip()),
    }


def clean_body(html: str, title: str = "", drop_leading_h1: bool = True, keep_h1: bool = False) -> str:
    """Remove the Shopify body's own H1 (the page already renders a title) and demote the rest to H2.

    keep_h1=True leaves the heading levels exactly as Shopify had them — the live product pages show
    the body's "Какво е X?" as the H1 and the product name as an H2.
    """
    if not html:
        return ""
    out = html.strip()
    if not keep_h1:
        if drop_leading_h1:
            m = re.match(r"\s*<h1[^>]*>.*?</h1>", out, flags=re.I | re.S)
            if m:
                out = out[m.end():].lstrip()
        out = re.sub(r"<h1(\s[^>]*)?>", "<h2>", out, flags=re.I)
        out = re.sub(r"</h1>", "</h2>", out, flags=re.I)
    # Shopify bodies can carry app scripts (hCaptcha, tracking) — they never belong in our HTML
    out = re.sub(r"<script\b.*?</script>", "", out, flags=re.I | re.S)
    out = re.sub(r"<script\b[^>]*/?>", "", out, flags=re.I)
    return out.strip()


# ---------- importers ----------
def import_collections() -> Dict[str, str]:
    rows = sheet("Custom Collections")
    groups = group_by(rows, "Handle")
    db.collections_cat.delete_many({})
    order_extra = 10
    imported = 0
    for handle, group in groups.items():
        top = group[0]
        published = str(top.get("Published")) in ("True", "true", "1")
        count = int(num(top.get("Products Count")))
        if not published:
            continue
        our_handle = ALL_HANDLE if handle == SHOPIFY_ALL else handle
        menu_title, sort_order = COLLECTION_META.get(our_handle, (top.get("Title") or our_handle, order_extra))
        if our_handle not in COLLECTION_META:
            order_extra += 1
        body = top.get("Body HTML") or ""
        db.collections_cat.insert_one({
            "id": str(uuid.uuid4()),
            "handle": our_handle,
            "link_key": link_key_for("collection", our_handle) or "",
            "title": (COLLECTION_META[ALL_HANDLE][0] if our_handle == ALL_HANDLE
                      else top.get("Title") or our_handle),
            "menu_title": menu_title,
            "menu_order": sort_order,
            "sort_order": sort_order,
            # published landing collections without products (SEO pages) stay reachable but out of the nav
            "nav_hidden": count == 0,
            # heading levels exactly as on purepeptide.bg: the body's "Какво е X?" is the page H1
            "description": clean_body(rewrite_body_images(body), top.get("Title") or "", keep_h1=True),
            **seo_pair(top, top.get("Title") or our_handle, body),
            "image": store_image(top.get("Image Src")),
            "shopify_id": str(top.get("ID") or ""),
            "translations": {},
            "created_at": now_utc(),
        })
        imported += 1
    log.info("collections imported: %d", imported)
    return {}


def import_products() -> None:
    from seed_data import PRODUCTS as SEED_PRODUCTS
    seed_specs = {p["handle"]: p.get("specs") or {} for p in SEED_PRODUCTS}
    rows = sheet("Products")
    groups = group_by(rows, "Handle")
    db.products.delete_many({})
    imported = 0
    for handle, group in groups.items():
        top = group[0]
        body = clean_body(rewrite_body_images(top.get("Body HTML")), top.get("Title") or handle)
        images: List[str] = []
        seen = set()
        for r in sorted(group, key=lambda x: num(x.get("Image Position"), 99)):
            src = r.get("Image Src")
            if src and src not in seen:
                seen.add(src)
                images.append(store_image(src))
        variants = []
        for r in group:
            name = r.get("Option1 Value")
            price = r.get("Variant Price")
            if not name or price is None:
                continue
            # the store sells through the Bulgarian market, where the "was" price lives
            compare = r.get("Variant Compare At Price")
            if compare is None:
                compare = r.get("Compare At Price / Bulgaria")
            variants.append({
                "name": str(name),
                "price_eur": round(num(price), 2),
                "compare_at_eur": round(num(compare), 2) if num(compare) > num(price) else 0.0,
                "stock": int(num(r.get("Variant Inventory Qty"))),
                "sku": r.get("Variant SKU") or "",
            })
        if not variants:
            variants = [{"name": "1 бр.", "price_eur": 0.0, "stock": 0, "sku": ""}]
        cols = []
        for c in str(top.get("Custom Collections") or "").split(","):
            c = c.strip()
            if not c:
                continue
            cols.append(ALL_HANDLE if c == SHOPIFY_ALL else c)
        tags = [t.strip() for t in str(top.get("Tags") or "").split(",") if t.strip()]
        db.products.insert_one({
            "id": str(uuid.uuid4()),
            "handle": handle,
            "title": top.get("Title") or handle,
            "subtitle": "",
            "description": body,
            "image": images[0] if images else "",
            "images": images,
            "variants": variants,
            "collections": cols,
            "tags": tags,
            "specs": extract_specs(top.get("Body HTML") or "") or seed_specs.get(handle, {}),
            **seo_pair(top, top.get("Title") or handle, top.get("Body HTML") or ""),
            "featured": False,
            "active": True,
            "shopify_status": str(top.get("Status") or ""),
            "shopify_id": str(top.get("ID") or ""),
            "translations": {},
            "created_at": parse_dt(top.get("Created At")),
        })
        imported += 1
    log.info("products imported: %d", imported)


SKIP_PAGE_HANDLES = {"ads-page", "homepage"}


def _page_upsert(slug: str, title: str, html: str, seo: Dict[str, str],
                 canonical_slug: Optional[str] = None) -> None:
    fields = {"title": title, "html": html, "faq_items": [], **seo, "updated_at": now_utc(),
              "link_key": link_key_for("page", canonical_slug or slug) or ""}
    fields["canonical_slug"] = canonical_slug or ""
    db.pages.update_one(
        {"slug": slug, "locale": "bg"},
        {"$set": fields, "$setOnInsert": {"id": str(uuid.uuid4()), "slug": slug, "locale": "bg"}},
        upsert=True,
    )


def import_pages() -> None:
    """Every page keeps its Shopify handle as the slug — that is what the storefront links to.

    PAGE_MAP handles are additionally published under the app's own slug (an alias carrying
    canonical_slug, so sitemaps skip it) to keep older internal links alive.
    """
    for handle, group in group_by(sheet("Pages"), "Handle").items():
        handle = str(handle or "").strip()
        if not handle or handle in SKIP_PAGE_HANDLES or handle.startswith("html-sitemap"):
            continue
        top = next((r for r in group if str(r.get("Body HTML") or "").strip()), group[0])
        body = str(top.get("Body HTML") or "")
        title = top.get("Title") or handle
        html = clean_body(rewrite_body_images(body), title)
        if not html.strip():
            log.info("page skipped (no content): %s", handle)
            continue
        seo = seo_pair(top, title, body)
        _page_upsert(handle, title, html, seo)
        alias = PAGE_MAP.get(handle)
        if alias and alias != handle:
            _page_upsert(alias, title, html, seo, canonical_slug=handle)
        log.info("page imported: %s%s", handle, f" (+ alias {alias})" if alias else "")


def import_articles() -> None:
    rows = sheet("Blog Posts")
    groups = group_by(rows, "Handle")
    product_handles = [(p["handle"], p["title"]) for p in db.products.find({}, {"handle": 1, "title": 1})]
    db.articles.delete_many({})
    imported = 0
    for handle, group in groups.items():
        top = group[0]
        body = top.get("Body HTML") or ""
        if not body.strip():
            continue
        title = top.get("Title") or handle
        linked = ""
        low = f"{handle} {title}".lower()
        for ph, ptitle in product_handles:
            core = re.sub(r"[^a-z0-9]+", "", ph.lower())
            key = re.sub(r"[^a-z0-9]+", "", str(ptitle).lower())
            if (len(core) > 5 and core in re.sub(r"[^a-z0-9]+", "", low)) or (len(key) > 5 and key in re.sub(r"[^a-z0-9]+", "", low)):
                linked = ph
                break
        db.articles.insert_one({
            "id": str(uuid.uuid4()),
            "handle": handle,
            "title": title,
            "excerpt": strip_html(top.get("Summary HTML") or body, 220),
            "body": clean_body(rewrite_body_images(body), title),
            "image": store_image(top.get("Image Src")),
            "author": top.get("Author") or "PurePeptide",
            "published": str(top.get("Published")) in ("True", "true", "1"),
            "product_handle": linked,
            **seo_pair(top, title, top.get("Summary HTML") or body),
            "published_at": parse_dt(top.get("Published At") or top.get("Created At")),
            "translations": {},
        })
        imported += 1
    log.info("articles imported: %d", imported)


def import_redirects() -> None:
    rows = sheet("Redirects")
    for r in rows:
        path = r.get("Path")
        target = r.get("Target")
        if not path:
            continue
        db.delisted_links.update_one(
            {"url": path},
            {
                "$set": {
                    "url": path,
                    "locale": "bg",
                    "reason": "Shopify 301 redirect",
                    "status": "redirected",
                    "replacement_url": target or "",
                    "notes": "Импортирано от Matrixify",
                    "updated_at": now_utc(),
                },
                "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now_utc(), "created_by": "matrixify-import"},
            },
            upsert=True,
        )
    log.info("redirects imported: %d", len(rows))


def import_discounts() -> None:
    rows = sheet("Discounts")
    codes = []
    for r in rows:
        code = r.get("Code")
        if not code:
            continue
        value_type = str(r.get("Value Type") or "").lower()
        codes.append({
            "code": str(code).strip().upper(),
            "type": "percent" if "percent" in value_type else "fixed",
            "value": round(abs(num(r.get("Value"))), 2),
            "min_subtotal": round(num(r.get("Minimum Value")), 2),
            "active": str(r.get("Status") or "").lower() == "active",
            "starts_at": parse_dt(r.get("Starts At")),
            "ends_at": parse_dt(r.get("Ends At")) if r.get("Ends At") else "",
            "used_count": int(num(r.get("Total Used Count"))),
            "title": r.get("Title") or "",
        })
    db.settings.update_one({"key": "site"}, {"$set": {"value.discount_codes": codes, "updated_at": now_utc()}})
    log.info("discount codes imported: %d (active: %d)", len(codes), sum(1 for c in codes if c["active"]))


def import_customers() -> None:
    rows = sheet("Customers")
    groups = group_by(rows, "ID")
    db.customers.delete_many({})
    docs = []
    for shopify_id, group in groups.items():
        top = group[0]
        email = top.get("Email")
        if not email:
            continue
        addr = next((r for r in group if r.get("Address Line 1") or r.get("Address City")), top)
        name = " ".join(x for x in [top.get("First Name"), top.get("Last Name")] if x).strip()
        docs.append({
            "id": str(uuid.uuid4()),
            "shopify_id": str(shopify_id),
            "email": str(email).strip().lower(),
            "name": name or (addr.get("Address First Name") or "") + " " + (addr.get("Address Last Name") or ""),
            "phone": top.get("Phone") or addr.get("Address Phone") or "",
            "language": top.get("Language") or "bg",
            "note": top.get("Note") or "",
            "tags": [t.strip() for t in str(top.get("Tags") or "").split(",") if t.strip()],
            "accepts_marketing": str(top.get("Email Marketing: Status") or "").lower() == "subscribed",
            "total_spent": round(num(top.get("Total Spent")), 2),
            "total_orders": int(num(top.get("Total Orders"))),
            "first_order_at": parse_dt(top.get("First Order: Processed At")) if top.get("First Order: Processed At") else "",
            "last_order_at": parse_dt(top.get("Last Order: Processed At")) if top.get("Last Order: Processed At") else "",
            "address": {
                "line1": addr.get("Address Line 1") or "",
                "line2": addr.get("Address Line 2") or "",
                "city": addr.get("Address City") or "",
                "zip": addr.get("Address Zip") or "",
                "country": addr.get("Address Country") or "",
            },
            "created_at": parse_dt(top.get("Created At")),
            "source": "shopify_import",
        })
    if docs:
        db.customers.insert_many(docs)
    log.info("customers imported: %d", len(docs))


def import_orders() -> None:
    rows = sheet("Orders")
    groups = group_by(rows, "ID")
    db.orders.delete_many({"source": "shopify_import"})
    docs = []
    for shopify_id, group in groups.items():
        top = group[0]
        line_items = []
        for r in group:
            if str(r.get("Line: Type") or "").lower() != "line item":
                continue
            title = r.get("Line: Title") or r.get("Line: Name")
            if not title:
                continue
        cur = str(top.get("Currency") or "EUR").upper().strip()
        fx = rate_for(cur)
        for r in group:
            if str(r.get("Line: Type") or "").lower() != "line item":
                continue
            title = r.get("Line: Title") or r.get("Line: Name")
            if not title:
                continue
            price = round(num(r.get("Line: Price")), 2)
            line_items.append({
                "product_handle": r.get("Line: Product Handle") or "",
                "title": title,
                "variant": r.get("Line: Variant Title") or "",
                "sku": r.get("Line: SKU") or "",
                "quantity": int(num(r.get("Line: Quantity"), 1)),
                "price_eur": round(price / fx, 2),
                "price_orig": price,
            })
        total = round(num(top.get("Price: Total")), 2)
        payment = str(top.get("Payment: Status") or "").lower()
        fulfillment = str(top.get("Order Fulfillment Status") or "").lower()
        status = "cancelled" if top.get("Cancelled At") else (
            "shipped" if "fulfilled" in fulfillment else ("paid" if payment == "paid" else "awaiting_payment")
        )
        tracking = next((r.get("Fulfillment: Tracking Number") for r in group if r.get("Fulfillment: Tracking Number")), "")
        docs.append({
            "id": str(uuid.uuid4()),
            "shopify_id": str(shopify_id),
            "order_number": str(top.get("Name") or f"SH-{shopify_id}"),
            "customer_info": {
                "name": " ".join(x for x in [top.get("Customer: First Name"), top.get("Customer: Last Name")] if x).strip()
                        or top.get("Shipping: Name") or top.get("Billing: Name") or "",
                "email": str(top.get("Email") or top.get("Customer: Email") or "").lower(),
                "phone": top.get("Phone") or top.get("Shipping: Phone") or "",
                "address": top.get("Shipping: Address 1") or top.get("Billing: Address 1") or "",
                "city": top.get("Shipping: City") or top.get("Billing: City") or "",
                "zip": top.get("Shipping: Zip") or top.get("Billing: Zip") or "",
                "country": top.get("Shipping: Country") or "Bulgaria",
            },
            "line_items": line_items,
            "subtotal_eur": round(num(top.get("Price: Subtotal")) / fx, 2),
            "discount_eur": round(num(top.get("Price: Total Discount")) / fx, 2),
            "shipping_eur": round(num(top.get("Price: Total Shipping")) / fx, 2),
            "total_eur": round(total / fx, 2),
            "subtotal_orig": round(num(top.get("Price: Subtotal")), 2),
            "discount_orig": round(num(top.get("Price: Total Discount")), 2),
            "shipping_orig": round(num(top.get("Price: Total Shipping")), 2),
            "total_orig": total,
            "currency": cur,
            "currency_rate": fx,
            "currency_normalized": True,
            "payment_method": next((r.get("Transaction: Gateway") for r in group if r.get("Transaction: Gateway")), ""),
            "payment_status": "paid" if payment == "paid" else payment or "pending",
            "status": status,
            "tracking_number": tracking or "",
            "note": top.get("Note") or "",
            "created_at": parse_dt(top.get("Processed At") or top.get("Created At")),
            "source": "shopify_import",
        })
    if docs:
        for i in range(0, len(docs), 500):
            db.orders.insert_many(docs[i:i + 500], ordered=False)
    log.info("orders imported: %d", len(docs))


def backfill_customer_spend() -> None:
    """Recompute spend per customer from the imported orders (spending history)."""
    pipeline = [
        {"$match": {"source": "shopify_import", "status": {"$ne": "cancelled"}}},
        {"$group": {
            "_id": "$customer_info.email",
            "orders": {"$sum": 1},
            "spent": {"$sum": "$total_eur"},
            "last": {"$max": "$created_at"},
            "first": {"$min": "$created_at"},
        }},
    ]
    updated = 0
    for row in db.orders.aggregate(pipeline):
        email = row["_id"]
        if not email:
            continue
        res = db.customers.update_one(
            {"email": email},
            {"$set": {
                "total_orders": row["orders"],
                "total_spent": round(row["spent"], 2),
                "first_order_at": row["first"],
                "last_order_at": row["last"],
            }},
        )
        if res.matched_count == 0:
            db.customers.insert_one({
                "id": str(uuid.uuid4()),
                "email": email,
                "name": "",
                "phone": "",
                "total_orders": row["orders"],
                "total_spent": round(row["spent"], 2),
                "first_order_at": row["first"],
                "last_order_at": row["last"],
                "created_at": row["first"],
                "source": "shopify_import",
            })
        updated += 1
    log.info("customer spend backfilled: %d", updated)


STEPS = {
    "collections": import_collections,
    "products": import_products,
    "pages": import_pages,
    "articles": import_articles,
    "redirects": import_redirects,
    "discounts": import_discounts,
    "customers": import_customers,
    "orders": import_orders,
    "spend": backfill_customer_spend,
}


if __name__ == "__main__":
    only = None
    xlsx_path = XLSX
    for i, arg in enumerate(sys.argv):
        if arg == "--only" and i + 1 < len(sys.argv):
            only = [s.strip() for s in sys.argv[i + 1].split(",")]
        if arg == "--file" and i + 1 < len(sys.argv):
            xlsx_path = Path(sys.argv[i + 1])
    WB = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    available = set(WB.sheetnames)
    storage.init_storage()
    for name, fn in STEPS.items():
        if only and name not in only:
            continue
        sheet_name = {"collections": "Custom Collections", "products": "Products", "pages": "Pages",
                      "articles": "Blog Posts", "redirects": "Redirects", "discounts": "Discounts",
                      "customers": "Customers", "orders": "Orders"}.get(name)
        if sheet_name and sheet_name not in available:
            log.info("=== %s skipped (no '%s' sheet in file)", name, sheet_name)
            continue
        log.info("=== %s", name)
        fn()
    db.settings.update_one(
        {"key": "site"},
        {"$set": {"value.catalog_imported": True, "value.catalog_imported_at": now_utc()}},
    )
    log.info("import finished")
