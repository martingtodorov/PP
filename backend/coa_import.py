"""Chemical-analysis (COA) files from the Shopify export -> product galleries.

The Matrixify export keeps the lab report of a product in the metafield
`custom.chemical_analysis` (a file reference, e.g. `Test_Report_Sermorelin.png`) and the download
link of that file in the `Files` sheet. Nothing in the storefront ever read it, so the reports had
to be uploaded by hand. `pairs()` resolves metafield -> CDN link per product handle.
"""
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

XLSX = Path(__file__).parent / "data" / "matrixify-export.xlsx"
META_COLS = (
    "Metafield: custom.chemical_analysis [file_reference]",
    "Metafield: custom.chemical_analysis [string]",
)


def pairs() -> List[Dict[str, Any]]:
    """[{handle, filename, url}] for every product that has a chemical-analysis file."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    links: Dict[str, str] = {}
    if "Files" in wb.sheetnames:
        rows = wb["Files"].iter_rows(values_only=True)
        hdr = list(next(rows))
        name_i, link_i = hdr.index("File Name"), hdr.index("Link")
        for r in rows:
            name, link = r[name_i], r[link_i]
            if name and link and str(name) not in links:
                links[str(name).strip()] = str(link).strip()

    rows = wb["Products"].iter_rows(values_only=True)
    hdr = list(next(rows))
    handle_i = hdr.index("Handle")
    meta_i = [hdr.index(c) for c in META_COLS if c in hdr]
    out: List[Dict[str, Any]] = []
    seen = set()
    for r in rows:
        handle = str(r[handle_i] or "").strip()
        filename = next((str(r[i]).strip() for i in meta_i if r[i]), "")
        if not handle or not filename or handle in seen:
            continue
        seen.add(handle)
        out.append({"handle": handle, "filename": filename, "url": links.get(filename, "")})
    return out
